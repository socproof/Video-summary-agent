import re
import yaml
import json
from pathlib import Path
from typing import List, Optional
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from langchain_ollama import OllamaLLM
from langchain_core.prompts import PromptTemplate


@dataclass
class TranscriptLine:
    start_time: float
    end_time: float
    text: str


@dataclass
class Topic:
    number: int
    start_time: str
    summary: str


class Config:
    """Управление конфигурацией"""
    def __init__(self, config_path: str = "config.yaml"):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                self.config = yaml.safe_load(f)
        except FileNotFoundError:
            print(f"⚠️  Конфиг не найден, используются настройки по умолчанию")
            self.config = self._default_config()

    def _default_config(self):
        return {
            'model': {'name': 'mistral', 'temperature': 0.3, 'num_ctx': 8192, 'timeout': 60},
            'processing': {
                'pause_threshold': 8,
                'min_topic_duration': 30,
                'max_topic_duration': 300,
                'use_llm': True
            },
            'paths': {'input_dir': 'transcripts', 'output_dir': 'results'},
            'output': {'format': 'xlsx'}
        }

    def get(self, *keys, default=None):
        value = self.config
        for key in keys:
            if isinstance(value, dict):
                value = value.get(key)
                if value is None:
                    return default
            else:
                return default
        return value


class TranscriptAnalyzer:
    """Гибридный анализатор: сегментация по паузам + батч-анализ через LLM"""

    def __init__(self, config_path: str = "config.yaml"):
        self.config = Config(config_path)
        self.pause_threshold = self.config.get('processing', 'pause_threshold', default=8)
        self.min_duration = self.config.get('processing', 'min_topic_duration', default=30)
        self.max_duration = self.config.get('processing', 'max_topic_duration', default=300)
        self.use_llm = self.config.get('processing', 'use_llm', default=True)

        print(f"🤖 Инициализация LLM...")
        self._init_llm()
        self._init_prompt()

    def _init_llm(self):
        """Инициализация LLM"""
        if not self.use_llm:
            self.llm = None
            return

        try:
            model_name = self.config.get('model', 'name', default='mistral')
            base_url = self.config.get('model', 'base_url', default='http://localhost:11434')

            print(f"  Модель: {model_name}")

            self.llm = OllamaLLM(
                model=model_name,
                base_url=base_url,
                temperature=self.config.get('model', 'temperature', default=0.3),
                num_ctx=self.config.get('model', 'num_ctx', default=8192)
            )
            print(f"  ✓ LLM готов")
        except Exception as e:
            print(f"  ✗ Ошибка LLM: {e}")
            self.llm = None

    def _init_prompt(self):
        """Инициализация промпта для батч-анализа"""
        self.batch_prompt = PromptTemplate(
            input_variables=["segments"],
            template="""Для каждого пронумерованного сегмента транскрипции сформулируй ОДНУ ГЛАВНУЮ ТЕМУ или КЛЮЧЕВОЙ ВОПРОС.

Требования к каждому описанию:
- Одно короткое предложение (5-15 слов)
- Суть темы или главный вопрос
- БЕЗ вводных слов "обсуждается", "говорится", "речь идёт"
- Конкретно и понятно

СЕГМЕНТЫ:
{segments}

Верни ТОЛЬКО JSON массив в формате:
[
  {{"segment": 1, "topic": "Краткое описание темы 1"}},
  {{"segment": 2, "topic": "Краткое описание темы 2"}},
  ...
]

JSON:"""
        )

    def parse_transcript(self, file_path: Path) -> List[TranscriptLine]:
        """Парсинг файла транскрипции"""
        print(f"📄 Парсинг файла...")
        lines = []

        patterns = [
            r'\[.*?\]\s*(\d+\.?\d*)-(\d+\.?\d*):\s*(.*)',
            r'(\d+\.?\d*)-(\d+\.?\d*):\s*(.*)',
            r'\[(\d+\.?\d*)\s*-\s*(\d+\.?\d*)\]\s*(.*)',
        ]

        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue

                for pattern in patterns:
                    match = re.match(pattern, line)
                    if match:
                        groups = match.groups()
                        lines.append(TranscriptLine(
                            start_time=float(groups[0]),
                            end_time=float(groups[1]),
                            text=groups[2].strip()
                        ))
                        break

        print(f"  ✓ Прочитано строк: {len(lines)}")
        return lines

    def segment_by_pauses(self, lines: List[TranscriptLine]) -> List[List[TranscriptLine]]:
        """Разбивка на сегменты по паузам с учётом min/max длительности"""
        if not lines:
            return []

        raw_segments = []
        current_segment = [lines[0]]

        # Первичная разбивка по паузам
        for i in range(1, len(lines)):
            pause = lines[i].start_time - lines[i-1].end_time
            current_duration = lines[i-1].end_time - current_segment[0].start_time

            # Разбиваем если пауза большая ИЛИ превышена макс. длительность
            if pause > self.pause_threshold or current_duration > self.max_duration:
                raw_segments.append(current_segment)
                current_segment = [lines[i]]
            else:
                current_segment.append(lines[i])

        if current_segment:
            raw_segments.append(current_segment)

        # Объединяем короткие сегменты
        final_segments = []
        i = 0
        while i < len(raw_segments):
            segment = raw_segments[i]
            duration = segment[-1].end_time - segment[0].start_time

            # Если сегмент слишком короткий, пробуем объединить со следующим
            if duration < self.min_duration and i + 1 < len(raw_segments):
                next_segment = raw_segments[i + 1]
                combined_duration = next_segment[-1].end_time - segment[0].start_time

                # Объединяем если это не превысит макс. длительность
                if combined_duration <= self.max_duration:
                    segment.extend(next_segment)
                    i += 2
                else:
                    final_segments.append(segment)
                    i += 1
            else:
                final_segments.append(segment)
                i += 1

        print(f"  ✓ Найдено сегментов: {len(final_segments)}")
        print(f"    (min: {self.min_duration}s, max: {self.max_duration}s, пауза: {self.pause_threshold}s)")
        return final_segments

    def batch_analyze_segments(self, segments: List[List[TranscriptLine]]) -> List[str]:
        """Батч-анализ всех сегментов одним запросом к LLM"""
        if not self.llm:
            print("  ✗ LLM не инициализирован, используем fallback")
            return [self._fallback_summary(seg) for seg in segments]

        print(f"  💭 Батч-анализ {len(segments)} сегментов через LLM...")

        # Форматируем все сегменты для одного запроса
        formatted_segments = []
        for i, segment in enumerate(segments, 1):
            text = ' '.join([line.text for line in segment])
            # Ограничиваем каждый сегмент для экономии токенов
            if len(text) > 800:
                text = text[:800] + "..."

            start_time = self._format_time(segment[0].start_time)
            formatted_segments.append(f"Сегмент {i} [{start_time}]:\n{text}")

        all_segments_text = "\n\n".join(formatted_segments)

        # Проверяем размер
        char_count = len(all_segments_text)
        estimated_tokens = char_count // 4
        print(f"    Размер батча: {char_count} символов (~{estimated_tokens} токенов)")

        if estimated_tokens > self.config.get('model', 'num_ctx', default=8192) * 0.7:
            print(f"    ⚠️  Батч большой, возможна обрезка. Рекомендуется увеличить num_ctx")

        try:
            prompt = self.batch_prompt.format(segments=all_segments_text)

            print(f"    Ожидание ответа от LLM...")
            import time
            start = time.time()

            response = self.llm.invoke(prompt)

            elapsed = time.time() - start
            print(f"    ✓ Получен ответ за {elapsed:.1f}s")

            # Парсим JSON
            summaries_data = self._parse_json_response(response)

            if not summaries_data or len(summaries_data) != len(segments):
                print(f"    ⚠️  Количество тем не совпадает ({len(summaries_data) if summaries_data else 0} != {len(segments)})")
                print(f"    Используем fallback для недостающих")

                summaries = []
                for i in range(len(segments)):
                    if summaries_data and i < len(summaries_data):
                        summary = summaries_data[i].get('topic', '')
                        if summary and len(summary) > 5:
                            summaries.append(summary)
                        else:
                            summaries.append(self._fallback_summary(segments[i]))
                    else:
                        summaries.append(self._fallback_summary(segments[i]))

                return summaries

            # Извлекаем только тексты тем
            summaries = [item.get('topic', 'Тема не определена') for item in summaries_data]
            print(f"    ✓ Успешно обработано тем: {len(summaries)}")

            return summaries

        except Exception as e:
            print(f"    ✗ Ошибка LLM: {e}")
            import traceback
            traceback.print_exc()
            print(f"    Используем fallback для всех сегментов")
            return [self._fallback_summary(seg) for seg in segments]

    def _fallback_summary(self, segment: List[TranscriptLine]) -> str:
        """Простое саммари без LLM"""
        text = ' '.join([line.text for line in segment])
        words = [w for w in text.split() if len(w) > 3][:12]
        return ' '.join(words) if words else "Тема не определена"

    def _parse_json_response(self, response: str) -> Optional[List[dict]]:
        """Парсинг JSON из ответа LLM"""
        try:
            # Убираем markdown если есть
            response = re.sub(r'```json\s*', '', response)
            response = re.sub(r'```\s*', '', response)
            response = response.strip()

            # Ищем JSON массив
            start = response.find('[')
            end = response.rfind(']') + 1

            if start != -1 and end > start:
                json_str = response[start:end]
                parsed = json.loads(json_str)
                return parsed

            # Пробуем парсить весь ответ
            return json.loads(response)

        except Exception as e:
            print(f"    ✗ Ошибка парсинга JSON: {e}")
            print(f"    Ответ LLM (первые 300 символов): {response[:300]}")
            return None

    def analyze(self, lines: List[TranscriptLine]) -> List[Topic]:
        """Основной метод анализа"""
        if not lines:
            return []

        print(f"  📊 Строк: {len(lines)}")
        total_duration = lines[-1].end_time - lines[0].start_time
        print(f"  ⏱️  Длительность: {total_duration/60:.1f} минут")

        print(f"\n  🔍 Этап 1: Сегментация по паузам...")
        segments = self.segment_by_pauses(lines)

        if not segments:
            print("  ✗ Не удалось выделить сегменты")
            return []

        print(f"\n  🔍 Этап 2: Генерация описаний...")
        summaries = self.batch_analyze_segments(segments)

        # Формируем итоговые темы
        topics = []
        for i, (segment, summary) in enumerate(zip(segments, summaries), 1):
            topics.append(Topic(
                number=i,
                start_time=self._format_time(segment[0].start_time),
                summary=summary
            ))

        return topics

    def save_xlsx(self, topics: List[Topic], output_path: Path):
        """Сохранение в XLSX"""
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Темы"

        # Заголовки
        headers = ["Время", "Тема/Вопрос"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Данные
        for row, topic in enumerate(topics, 2):
            ws.cell(row, 1, topic.start_time).alignment = Alignment(horizontal='center')
            ws.cell(row, 2, topic.summary).alignment = Alignment(wrap_text=True, vertical='top')

        # Ширина колонок
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 80

        # Высота строк
        ws.row_dimensions[1].height = 25
        for row in range(2, len(topics) + 2):
            ws.row_dimensions[row].height = 40

        wb.save(output_path)

    def process_file(self, input_file: Path) -> Optional[Path]:
        """Обработка файла"""
        print(f"\n{'='*60}")
        print(f"📂 Обработка: {input_file.name}")
        print(f"{'='*60}")

        lines = self.parse_transcript(input_file)

        if not lines:
            print("  ✗ Не удалось распарсить файл")
            return None

        topics = self.analyze(lines)

        if not topics:
            print("  ✗ Темы не выделены")
            return None

        print(f"\n  ✅ Итого тем: {len(topics)}")

        output_dir = Path(self.config.get('paths', 'output_dir', default='results'))
        output_path = output_dir / f"{input_file.stem}_topics.xlsx"

        self.save_xlsx(topics, output_path)
        print(f"  💾 Сохранено: {output_path}\n")

        return output_path

    def _format_time(self, seconds: float) -> str:
        """Форматирование времени MM:SS"""
        m = int(seconds // 60)
        s = int(seconds % 60)
        return f"{m:02d}:{s:02d}"


def main():
    import sys

    if len(sys.argv) < 2:
        print("Использование:")
        print("  python transcript_analyzer.py <файл.txt>")
        return

    analyzer = TranscriptAnalyzer("config.yaml")
    input_file = Path(sys.argv[1])

    if not input_file.exists():
        print(f"Ошибка: файл {input_file} не найден")
        return

    analyzer.process_file(input_file)


if __name__ == "__main__":
    main()